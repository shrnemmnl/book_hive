from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from users.models import CustomUser, Order, OrderItem, Review, Wallet, Transaction, CustomerSupport # Import your user model
from .models import Genre, Product, Variant, ProductImage,Coupon,CouponUsage
from django.contrib.auth.decorators import user_passes_test
from django.views.decorators.cache import never_cache, cache_control
from functools import wraps
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import reverse
from django.shortcuts import  get_object_or_404
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import F,Q, Sum, Count, DecimalField
from django.db.models.functions import TruncDate
import base64
import re
from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.core.mail import send_mail
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO
import logging



logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(name)s - %(message)s')
logger = logging.getLogger(__name__)  # Create logger





# Create your views here.
@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
def admin_signin(request):

    if request.user.is_authenticated:
        return redirect('admin_dashboard')
    
    if request.method == 'POST':
        email=request.POST.get('email', "").strip()
        password=request.POST.get('password', "").strip()

        isuser=authenticate(request, email=email, password=password)
        

        if isuser is not None and isuser.is_superuser:
            login(request, isuser)
            messages.success(request, 'You are successfully logged in!')
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid credentials or insufficient permissions.')

    return render(request, 'admin/admin_signin.html')




@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
@login_required(login_url='admin_signin')  # Redirect to login page if not authenticated
def admin_dashboard(request):
    """Admin Dashboard with charts, best selling products and categories"""
    
    # Get filter from request
    filter_type = request.GET.get('filter', 'month')  # Default to monthly view
    
    now = timezone.now()
    
    # Filter orders based on selected period
    orders = Order.objects.filter(
        status__in=['pending', 'shipped', 'delivered'],
        is_active=True
    ).select_related('user')
    
    now = timezone.now()
    filter_label = "All Time"
    
    if filter_type == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        filter_label = "Today"
    elif filter_type == 'week':
        start_date = now - timedelta(days=7)
        filter_label = "Past 7 Days"
    elif filter_type == 'month':
        start_date = now - timedelta(days=30)
        filter_label = "Past 30 Days"
    elif filter_type == 'year':
        start_date = now - timedelta(days=365)
        filter_label = "Past Year"
    else:  # all
        start_date = None
        filter_label = "All Time"
    
    if start_date:
        orders = orders.filter(created_at__gte=start_date)
    
    # Calculate sales data for chart (daily data for the period)
    if filter_type in ['today', 'week', 'month', 'year']:
        # Get daily sales data
        daily_sales = orders.annotate(
            day=TruncDate('created_at')
        ).values('day').annotate(
            total=Sum('net_amount'),
            count=Count('id')
        ).order_by('day')
        
        chart_labels = [item['day'].strftime('%d %b') if item['day'] else '' for item in daily_sales]
        chart_revenue = [float(item['total']) for item in daily_sales]
        chart_orders = [item['count'] for item in daily_sales]
    else:
        # For "all time", show monthly data - grouping by year-month
        # Using TruncDate might not work perfectly for monthly, so we'll iterate through orders
        from collections import OrderedDict
        monthly_data = OrderedDict()
        for order in orders:
            month_key = order.created_at.strftime('%Y-%m')
            if month_key not in monthly_data:
                monthly_data[month_key] = {'total': 0, 'count': 0}
            monthly_data[month_key]['total'] += float(order.net_amount)
            monthly_data[month_key]['count'] += 1
        
        chart_labels = list(monthly_data.keys())
        chart_revenue = [monthly_data[k]['total'] for k in chart_labels]
        chart_orders = [monthly_data[k]['count'] for k in chart_labels]
    
    # Overall statistics
    total_stats = orders.aggregate(
        total_orders=Count('id'),
        total_revenue=Sum('net_amount')
    )
    
    # Calculate average order value
    total_orders_count = total_stats['total_orders'] or 0
    total_revenue_amount = total_stats['total_revenue'] or 0
    avg_order_value = (total_revenue_amount / total_orders_count) if total_orders_count > 0 else 0
    
    # Best selling products (top 10)
    best_selling_products = OrderItem.objects.filter(
        order__in=orders
    ).values(
        'product_variant__product__book_title',
        'product_variant__product__author'
    ).annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('discount_price'))
    ).order_by('-total_sold')[:10]
    
    # Best selling categories (genres) (top 10)
    best_selling_categories = OrderItem.objects.filter(
        order__in=orders
    ).values(
        'product_variant__product__genre__genre_name'
    ).annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('discount_price')),
        order_count=Count('order', distinct=True)
    ).order_by('-total_sold')[:10]
    
    # Convert to JSON-safe format for Chart.js
    import json
    
    context = {
        'filter_type': filter_type,
        'filter_label': filter_label,
        'total_orders': total_orders_count,
        'total_revenue': total_revenue_amount,
        'avg_order_value': avg_order_value,
        'chart_labels': json.dumps(chart_labels),
        'chart_revenue': json.dumps(chart_revenue),
        'chart_orders': json.dumps(chart_orders),
        'best_selling_products': best_selling_products,
        'best_selling_categories': best_selling_categories,
    }
    
    return render(request, 'admin/admin_dashboard.html', context)



    



@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
@login_required(login_url='admin_signin')  
def books(request):
    
    books = Product.objects.select_related('genre').order_by('-updated_at')

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(books, 5)
    books = paginator.get_page(page)


    return render(request, 'admin/books.html', {"books":books})



@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
@login_required(login_url='admin_signin')  
def admin_order(request):

    order_details=Order.objects.prefetch_related('order_items').order_by('-created_at')

    page = request.GET.get('page', 1)
    paginator = Paginator(order_details, 10)
    order_details = paginator.get_page(page)

    return render(request, 'admin/admin_order.html', {"order_details":order_details})



@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
@login_required(login_url='admin_signin')  
def update_order_item_status(request,order_item_id):
    status=request.POST.get('status')
    order_item=OrderItem.objects.get(id=order_item_id)
    order_item.status=status
    order_item.save()
    
    # Generate invoice if status is 'shipped' and invoice doesn't exist
    if status == 'shipped':
        from users.views import generate_invoice_for_order
        try:
            # Check if Invoice table exists by trying to access it
            from users.models import Invoice
            generate_invoice_for_order(order_item.order)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            # Log error but don't fail the status update
            logger.warning(f"Could not generate invoice for order {order_item.order.order_id}: {str(e)}. Make sure migrations are run.")

    return redirect('order')



def genre(request):

    genre= Genre.objects.all().order_by('-is_active')

    paginator = Paginator(genre, 5)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        genre_name = request.POST.get('genre', "").strip()
        is_offer = request.POST.get('is_offer') == 'on'
        offer_title = request.POST.get('offer_title', "").strip()
        discount_percentage = request.POST.get('discount_percentage', 0)
        
        if not genre_name:
            messages.error(request, 'Genre name is required.')
            return redirect('genre')
        
        if not re.match(r'^[A-Za-z\s]+$', genre_name):
            messages.error(request, 'Genre name must contain only alphabets and spaces.')
            return redirect('genre')
        
        if len(genre_name) > 255:
            messages.error(request, 'Genre name is too long (max 255 characters).')
            return redirect('genre')
        
        if Genre.objects.filter(genre_name=genre_name.lower()).exists():
            messages.error(request, 'Genre already exists.')
            return redirect('genre')
        
        try:
            discount_percentage = int(discount_percentage) if discount_percentage else 0
            if discount_percentage < 0:
                messages.error(request, 'Discount percentage cannot be negative.')
                return redirect('genre')
            if discount_percentage >= 100:
                messages.error(request, 'Discount percentage must be less than 100%.')
                return redirect('genre')
        except ValueError:
            discount_percentage = 0
        
        Genre.objects.create(
            genre_name=genre_name.lower(),
            is_active=True,
            is_offer=is_offer,
            offer_title=offer_title if is_offer else '',
            discount_percentage=discount_percentage if is_offer else 0
        )
        messages.success(request, 'Genre added successfully.')
        return redirect('genre')
    
    return render(request, 'admin/genre.html', {'page_obj':page_obj})




def change_genre_status(request):
    if request.method == 'POST':
        genre_id=request.POST.get('genre_id', "").strip()
        status = Genre.objects.get(id=genre_id)
        status.is_active=not status.is_active
        status.save()
        return redirect('genre')
    return render(request, 'admin/genre.html')




def genre_edit(request, genre_id):

    genre = Genre.objects.get(id=genre_id)

    if request.method == 'POST':
        name = request.POST.get('genre', "").strip()
        is_offer = request.POST.get('is_offer') == 'on'
        offer_title = request.POST.get('offer_title', "").strip()
        discount_percentage = request.POST.get('discount_percentage', 0)
        
        genre = Genre.objects.get(id=genre_id)
        
        if not name:
            messages.error(request, 'Genre name is required.')
            return redirect('genre_edit', genre_id=genre.id)
        
        # Validate genre name contains only alphabets (and spaces)
        if not re.match(r'^[A-Za-z\s]+$', name):
            messages.error(request, 'Genre name must contain only alphabets and spaces.')
            return redirect('genre_edit', genre_id=genre.id)
        
        if len(name) > 255:
            messages.error(request, 'Genre name is too long (max 255 characters).')
            return redirect('genre_edit', genre_id=genre.id)

        if Genre.objects.filter(genre_name=name.lower()).exclude(id=genre_id).exists():
            messages.error(request, 'Genre already exists.')
            return redirect('genre_edit', genre_id=genre.id)
        
        try:
            discount_percentage = int(discount_percentage) if discount_percentage else 0
            if discount_percentage < 0:
                messages.error(request, 'Discount percentage cannot be negative.')
                return redirect('genre_edit', genre_id=genre.id)
            if discount_percentage >= 100:
                messages.error(request, 'Discount percentage must be less than 100%.')
                return redirect('genre_edit', genre_id=genre.id)
        except ValueError:
            discount_percentage = 0
                
        genre.genre_name = name.lower()
        genre.is_offer = is_offer
        genre.offer_title = offer_title if is_offer else ''
        genre.discount_percentage = discount_percentage if is_offer else 0
        genre.save()
        messages.success(request, 'Genre Edited Successfully.')
        return redirect('genre')
 
    return render(request, 'admin/genre_edit.html', {'genres':genre})



def genre_search(request):
    
    
    search_query=request.POST.get('genre_search', "").strip()
    genre = Genre.objects.filter(genre_name__istartswith=search_query)

    paginator = Paginator(genre, 5)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
        
    
    return render(request, 'admin/genre.html', {'page_obj':page_obj})




def add_new_book(request):
    if request.method == 'POST':
        book_name = request.POST.get('book_name', "").strip()
        author = request.POST.get('author', "").strip()
        genre_id = request.POST.get('genre_id', "").strip()
        description = request.POST.get('description', "").strip()
        image = request.FILES.get('image')

        error = {}
        is_valid = True

        # Validate book title
        if not book_name:
            error['book_name'] = "Book title is required."
            is_valid = False
        elif len(book_name) < 3:
            error['book_name'] = "Book title must be at least 3 characters long."
            is_valid = False
        elif len(book_name) > 200:
            error['book_name'] = "Book title must not exceed 200 characters."
            is_valid = False
        elif not re.match(r'^[a-zA-Z0-9\s.,!?;:\'"()-]+$', book_name):
            error['book_name'] = "Book title contains invalid characters. Only letters, numbers, spaces, and basic punctuation are allowed."
            is_valid = False
        elif Product.objects.filter(book_title=book_name).exists():
            error['book_name'] = "Book with this title already exists."
            is_valid = False

        # Validate author
        if not author:
            error['author'] = "Author name is required."
            is_valid = False
        elif len(author) < 2:
            error['author'] = "Author name must be at least 2 characters long."
            is_valid = False
        elif len(author) > 100:
            error['author'] = "Author name must not exceed 100 characters."
            is_valid = False
        elif not re.match(r'^[a-zA-Z\s\'-]+$', author):
            error['author'] = "Author name contains invalid characters. Only letters, spaces, hyphens, and apostrophes are allowed."
            is_valid = False

        # Validate genre
        genre = None
        if not genre_id:
            error['genre_id'] = "Genre is required."
            is_valid = False
        else:
            try:
                genre = Genre.objects.get(id=int(genre_id))
            except (Genre.DoesNotExist, ValueError):
                error['genre_id'] = "Invalid genre selected."
                is_valid = False
                genre = None

        # Validate image
        if not image:
            error['image'] = "Book image is required."
            is_valid = False
        else:
            valid_extensions = ["jpg", "jpeg", "png", "gif", "webp"]
            valid_mime_types = ["image/jpeg", "image/jpg", "image/png", "image/gif", "image/webp"]
            max_file_size = 5 * 1024 * 1024  # 5MB
            
            # Check file size
            if image.size > max_file_size:
                error['image'] = f"File size exceeds 5MB limit. Please choose a smaller file."
                is_valid = False
            
            # Check file extension
            file_extension = image.name.split(".")[-1].lower() if "." in image.name else ""
            if not file_extension or file_extension not in valid_extensions:
                error['image'] = "Please upload a valid image file (jpg, jpeg, png, gif, or webp formats only)."
                is_valid = False
            
            # Check MIME type
            if is_valid:
                file_mime_type = image.content_type.lower() if hasattr(image, 'content_type') and image.content_type else ""
                if not file_mime_type or file_mime_type not in valid_mime_types:
                    error['image'] = "Invalid file type detected. Please upload a valid image file (jpg, jpeg, png, gif, or webp formats only)."
                    is_valid = False

        # Validate description
        if not description:
            error['description'] = "Book description is required."
            is_valid = False

        if not is_valid:
            all_genres = Genre.objects.all()
            return render(request, 'admin/add_new_book.html', {'all_genres': all_genres, 'error': error})

        # All validations passed, create the book
        Product.objects.create(
            book_title=book_name,
            author=author,
            genre=genre,
            is_active=True,
            description=description,
            image=image
        )
        
        messages.success(request, 'Book added successfully!')
        return render(request, 'admin/add_new_book.html', {'redirect': True, 'redirect_url': 'books'})

    all_genres = Genre.objects.all()   
    return render(request, 'admin/add_new_book.html', {'all_genres': all_genres})



def book_edit(request, book_id):
    
    book = Product.objects.get(id=book_id)
    genres = Genre.objects.all()
    error = {}

    if request.method == 'POST':
        book_name = request.POST.get('book_title', "").strip()
        author = request.POST.get('book_author', "").strip()
        description = request.POST.get('book_description', "").strip()
        genre_id = request.POST.get('genre_id', "").strip()
        image = request.FILES.get('image')
        # Offer Fields
        is_offer = request.POST.get('is_offer') == 'on'
        offer_title = request.POST.get('offer_title', "").strip()
        discount_percentage = request.POST.get('discount_percentage', "").strip()

        is_valid = True

        # Validate book title
        if not book_name:
            error['book_title'] = "Book title is required."
            is_valid = False
        elif len(book_name) < 3:
            error['book_title'] = "Book title must be at least 3 characters long."
            is_valid = False
        elif len(book_name) > 200:
            error['book_title'] = "Book title must not exceed 200 characters."
            is_valid = False
        elif not re.match(r'^[a-zA-Z0-9\s.,!?;:\'"()-]+$', book_name):
            error['book_title'] = "Book title contains invalid characters. Only letters, numbers, spaces, and basic punctuation are allowed."
            is_valid = False

        # Validate author
        if not author:
            error['book_author'] = "Author name is required."
            is_valid = False
        elif len(author) < 2:
            error['book_author'] = "Author name must be at least 2 characters long."
            is_valid = False
        elif len(author) > 100:
            error['book_author'] = "Author name must not exceed 100 characters."
            is_valid = False
        elif not re.match(r'^[a-zA-Z\s\'-]+$', author):
            error['book_author'] = "Author name contains invalid characters. Only letters, spaces, hyphens, and apostrophes are allowed."
            is_valid = False

        # Validate genre
        genre = None
        if not genre_id:
            error['genre_id'] = "Genre is required."
            is_valid = False
        else:
            try:
                genre = Genre.objects.get(id=int(genre_id))
            except (Genre.DoesNotExist, ValueError):
                error['genre_id'] = "Invalid genre selected."
                is_valid = False
                genre = None

        # Validate image (optional for edit, but if provided, must be valid)
        if image:
            valid_extensions = ["jpg", "jpeg", "png", "gif", "webp"]
            valid_mime_types = ["image/jpeg", "image/jpg", "image/png", "image/gif", "image/webp"]
            max_file_size = 5 * 1024 * 1024  # 5MB
            
            # Check file size
            if image.size > max_file_size:
                error['valid_image'] = f"File size exceeds 5MB limit. Please choose a smaller file."
                is_valid = False
            
            # Check file extension
            if is_valid:
                file_extension = image.name.split(".")[-1].lower() if "." in image.name else ""
                if not file_extension or file_extension not in valid_extensions:
                    error['valid_image'] = "Please upload a valid image file (jpg, jpeg, png, gif, or webp formats only)."
                    is_valid = False
            
            # Check MIME type
            if is_valid:
                file_mime_type = image.content_type.lower() if hasattr(image, 'content_type') and image.content_type else ""
                if not file_mime_type or file_mime_type not in valid_mime_types:
                    error['valid_image'] = "Invalid file type detected. Please upload a valid image file (jpg, jpeg, png, gif, or webp formats only)."
                    is_valid = False

        # Validate description
        if not description:
            error['book_description'] = "Book description is required."
            is_valid = False

        # Discount percentage validation
        if is_offer:
            if not discount_percentage:
                error['discount_percentage'] = "Discount percentage is required when offer is active."
                is_valid = False
            else:
                try:
                    discount_percentage = float(discount_percentage)
                    if discount_percentage < 0:
                        error['discount_percentage'] = "Discount percentage cannot be negative."
                        is_valid = False
                    elif discount_percentage >= 100:
                        error['discount_percentage'] = "Discount percentage must be less than 100%."
                        is_valid = False
                except (ValueError, AttributeError):
                    error['discount_percentage'] = "Discount percentage must be a valid number."
                    is_valid = False
        else:
            discount_percentage = 0

        if not is_valid:
            return render(request, 'admin/book_edit.html', {'book': book, 'genres': genres, 'error': error})

        # All validations passed, update the book
        book.book_title = book_name
        book.author = author
        book.description = description
        if genre:
            book.genre = genre
        if image:
            book.image = image
        book.is_offer = is_offer
        book.offer_title = offer_title if is_offer else ''
        book.discount_percentage = discount_percentage if is_offer else 0

        book.save()
        messages.success(request, f'{book.book_title} Updated successfully!')
        return redirect('books')  

    return render(request, 'admin/book_edit.html', {'book': book, 'genres': genres, 'error': error})





def book_delete(request):
    if request.method == 'POST':
        book_id = request.POST.get('book_delete', "").strip()
        status = Product.objects.get(id=book_id)
        status.is_active=not status.is_active
        status.save()
        return redirect('books')
        
    return render(request, 'admin/books.html',{'status':status})




def view_variant(request, book_id):

    variants = Variant.objects.filter(product=book_id).prefetch_related('productimage_set')

    first_variant=book_id

    return render(request, 'admin/view_variant.html', {
        'variants': variants,
        'first_variant':first_variant
    })


def variant_delete(request):
    if request.method == 'POST':
        variant_id = request.POST.get('variant_id', "").strip()  #comes from html

        variant = Variant.objects.get(id=variant_id)
        variant.is_active=not variant.is_active
        variant.save()
        return redirect(f'view_variant/{variant.product.id}')
        
    return render(request, 'admin/view_variant.html')



def add_variant(request, book_id):
    if request.method == 'POST':
        publisher = request.POST.get('publisher', "").strip()
        published_date = request.POST.get('published_date', "").strip()
        stock = request.POST.get('stock', "").strip()
        language = request.POST.get('language', "").strip()
        price = request.POST.get('price', "").strip()
        page = request.POST.get('page', "").strip()
        # Handling file uploads correctly
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        image3 = request.FILES.get('image3')

        error = {}
        is_valid = True

        try:
            product = Product.objects.get(id=book_id)  # Fetch Product instance
        except Product.DoesNotExist:
            messages.error(request, "Product not found")
            return redirect('books')
        
        # Validate publisher
        if not publisher:
            error['publisher'] = "Publisher is required."
            is_valid = False

        # Validate published date
        if not published_date:
            error['published_date'] = "Published date is required."
            is_valid = False
        else:
            try:
                pub_date = datetime.strptime(published_date, "%Y-%m-%d").date()
                today = timezone.now().date()
                if pub_date > today:
                    error['published_date'] = "Published date cannot be in the future."
                    is_valid = False
                # Check if date is too old (e.g., before 1900)
                min_date = datetime(1900, 1, 1).date()
                if pub_date < min_date:
                    error['published_date'] = "Published date must be after 1900."
                    is_valid = False
            except ValueError:
                error['published_date'] = "Invalid published date format."
                is_valid = False

        # Validate price
        if not price:
            error['price'] = "Price is required."
            is_valid = False
        else:
            try:
                price = float(price)
                if price <= 0:
                    error['price'] = "Price must be a positive number greater than 0."
                    is_valid = False
                elif price < 1:
                    error['price'] = "Price must be at least ₹1."
                    is_valid = False
                elif price > 100000:
                    error['price'] = "Price cannot exceed ₹1,00,000."
                    is_valid = False
            except ValueError:
                error['price'] = "Price must be a valid number."
                is_valid = False

        # Validate page count
        if not page:
            error['page'] = "Page count is required."
            is_valid = False
        else:
            try:
                page = int(page)
                if page < 1:
                    error['page'] = "Page count must be a positive integer (minimum 1)."
                    is_valid = False
                elif page > 10000:
                    error['page'] = "Page count cannot exceed 10,000."
                    is_valid = False
            except ValueError:
                error['page'] = "Page count must be a valid integer."
                is_valid = False

        # Validate stock (available quantity)
        if not stock:
            error['stock'] = "Available quantity is required."
            is_valid = False
        else:
            try:
                stock = int(stock)
                if stock < 0:
                    error['stock'] = "Available quantity must be a non-negative integer (minimum 0)."
                    is_valid = False
                elif stock > 100000:
                    error['stock'] = "Available quantity cannot exceed 1,00,000."
                    is_valid = False
            except ValueError:
                error['stock'] = "Available quantity must be a valid integer."
                is_valid = False

        # Validate language
        if not language:
            error['language'] = "Language is required."
            is_valid = False

        # Validate images
        valid_extensions = ["jpg", "jpeg", "png", "gif", "webp"]
        valid_mime_types = ["image/jpeg", "image/jpg", "image/png", "image/gif", "image/webp"]
        max_file_size = 5 * 1024 * 1024  # 5MB

        if not image1:
            error['image1'] = "Image 1 is required."
            is_valid = False
        else:
            if image1.size > max_file_size:
                error['image1'] = f"Image 1 exceeds 5MB limit. Please choose a smaller file."
                is_valid = False
            if is_valid:
                file_extension = image1.name.split(".")[-1].lower() if "." in image1.name else ""
                if not file_extension or file_extension not in valid_extensions:
                    error['image1'] = "Image 1: Invalid file type. Please upload images only (jpg, jpeg, png, gif, or webp formats)."
                    is_valid = False
            if is_valid:
                file_mime_type = image1.content_type.lower() if hasattr(image1, 'content_type') and image1.content_type else ""
                if not file_mime_type or file_mime_type not in valid_mime_types:
                    error['image1'] = "Image 1: Invalid file type. Please upload images only (jpg, jpeg, png, gif, or webp formats)."
                    is_valid = False

        if not image2:
            error['image2'] = "Image 2 is required."
            is_valid = False
        else:
            if image2.size > max_file_size:
                error['image2'] = f"Image 2 exceeds 5MB limit. Please choose a smaller file."
                is_valid = False
            if is_valid:
                file_extension = image2.name.split(".")[-1].lower() if "." in image2.name else ""
                if not file_extension or file_extension not in valid_extensions:
                    error['image2'] = "Image 2: Invalid file type. Please upload images only (jpg, jpeg, png, gif, or webp formats)."
                    is_valid = False
            if is_valid:
                file_mime_type = image2.content_type.lower() if hasattr(image2, 'content_type') and image2.content_type else ""
                if not file_mime_type or file_mime_type not in valid_mime_types:
                    error['image2'] = "Image 2: Invalid file type. Please upload images only (jpg, jpeg, png, gif, or webp formats)."
                    is_valid = False

        if not image3:
            error['image3'] = "Image 3 is required."
            is_valid = False
        else:
            if image3.size > max_file_size:
                error['image3'] = f"Image 3 exceeds 5MB limit. Please choose a smaller file."
                is_valid = False
            if is_valid:
                file_extension = image3.name.split(".")[-1].lower() if "." in image3.name else ""
                if not file_extension or file_extension not in valid_extensions:
                    error['image3'] = "Image 3: Invalid file type. Please upload images only (jpg, jpeg, png, gif, or webp formats)."
                    is_valid = False
            if is_valid:
                file_mime_type = image3.content_type.lower() if hasattr(image3, 'content_type') and image3.content_type else ""
                if not file_mime_type or file_mime_type not in valid_mime_types:
                    error['image3'] = "Image 3: Invalid file type. Please upload images only (jpg, jpeg, png, gif, or webp formats)."
                    is_valid = False

        if not is_valid:
            return render(request, 'admin/add_variant.html', {'book_id': book_id, 'error': error})
        
        # All validations passed, create the variant
        new_variant = Variant.objects.create(
            product=product,
            publisher=publisher,
            published_date=published_date,
            available_quantity=stock,
            language=language,
            page=page,
            price=price,
        )
        
        # Save images
        ProductImage.objects.create(
            variant=new_variant,
            image1=image1,
            image2=image2,
            image3=image3
        )

        messages.success(request, 'Variant Added successfully!')
        return render(request, 'admin/add_variant.html', {
            'redirect': True,
            'redirect_url': reverse('view_variant', args=[book_id])
        })

    return render(request, 'admin/add_variant.html', {'book_id': book_id})




@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
def admin_logout(request):

    logout(request)  # Clears the session and logs out the user
    request.session.flush()  # Ensures all session data is removed

    return redirect('admin_signin')



@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
@login_required(login_url='admin_signin')  
def customer_details(request):
    users = CustomUser.objects.all().exclude(is_superuser = True).annotate(
        query_count=Count('support_queries')
    ).order_by('-query_count', '-date_joined')

     # Create Paginator object with 10 customers per page
    paginator = Paginator(users, 10)  
    page_number = request.GET.get('page') # Get the current page number from request
    print(page_number) 
    page_obj = paginator.get_page(page_number)  # Get customers for the requested page
    print(page_obj)

    return render(request, 'admin/customer_details.html', {'users':page_obj})



def change_user_status(request):
    
    if request.method == 'POST':
        user_id=request.POST.get('user_id', "").strip()
        status = CustomUser.objects.get(id=user_id)
        status.is_active = not status.is_active
        status.save()

        return redirect('customer_details')
    
    return render(request, 'admin/customer_details.html')



def user_search(request):
    
    
    search_query=request.POST.get('user_search', "").strip()
    users = CustomUser.objects.filter(first_name__istartswith=search_query).exclude(is_superuser = True).annotate(
        query_count=Count('support_queries')
    ).order_by('-query_count', '-date_joined')

    # Create Paginator object with 10 customers per page
    paginator = Paginator(users, 10)  
    page_number = request.GET.get('page') # Get the current page number from request
    print(page_number) 
    page_obj = paginator.get_page(page_number)  # Get customers for the requested page
    print(page_obj)
        
    
    return render(request, 'admin/customer_details.html', {'users':page_obj})



def variant_edit(request, variant_id):
    error = {}    
    variant = get_object_or_404(Variant, id=variant_id)
    images = variant.productimage_set.first()  # set of 3 images per variant

    if request.method == 'POST':
        publisher = request.POST.get('publisher', "").strip()
        published_date = request.POST.get('published_date', "").strip()
        price = request.POST.get('price', "").strip()
        page = request.POST.get('page', "").strip()
        available_quantity = request.POST.get('stock', "").strip()
        language = request.POST.get('language', "").strip()

        # Get raw files
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        image3 = request.FILES.get('image3')

        # Get cropped base64 images
        image1_cropped = request.POST.get('image1_cropped')
        image2_cropped = request.POST.get('image2_cropped')
        image3_cropped = request.POST.get('image3_cropped')

        is_valid = True

        # --- Validations ---

        # 1️⃣ Published date
        try:           
            pub_date = datetime.strptime(published_date, "%Y-%m-%d").date()
            if pub_date > timezone.now().date():
                error['published_date'] = "Published date cannot be in the future."
                is_valid = False
        except ValueError:
            error['published_date'] = "Invalid published date format."
            is_valid = False

        # 2️⃣ Price
        if not price.replace('.', '', 1).isdigit():
            error['price'] = "Price must contain only digits."
            is_valid = False
        else:
            price = float(price)
            if price < 0:
                error['price'] = "Price cannot be negative."
                is_valid = False

        # 3️⃣ Page count
        if not page.isdigit() or int(page) <= 0:
            error['page'] = "Page count must be a positive integer."
            is_valid = False

        # 4️⃣ Available quantity
        if not available_quantity.isdigit() or int(available_quantity) <= 0:
            error['available_quantity'] = "Available quantity must be a positive integer."
            is_valid = False

        # 5️⃣ Image validation (raw files)
        valid_extensions = ["jpg", "jpeg", "png"]
        for img, name in zip([image1, image2, image3], ['Image 1', 'Image 2', 'Image 3']):
            if img:
                ext = img.name.split('.')[-1].lower()
                if ext not in valid_extensions:
                    error['valid_image'] = f"{name} must be in jpg, jpeg, or png format."
                    is_valid = False

        # 6️⃣ Cropped image validation
        for img_cropped, name in zip([image1_cropped, image2_cropped, image3_cropped],
                                     ['Image 1', 'Image 2', 'Image 3']):
            if img_cropped and not img_cropped.startswith("data:image"):
                error['valid_image'] = f"{name} has invalid cropped image format."
                is_valid = False

        # --- Save Variant ---
        if is_valid:
            variant.publisher = publisher
            variant.price = float(price)
            variant.page = int(page)
            variant.available_quantity = int(available_quantity)
            variant.language = language
            variant.save()

            # Handle ProductImage
            product_image, _ = ProductImage.objects.get_or_create(variant=variant)

            # Helper to save cropped/base64 image
            def save_cropped_image(cropped, raw, field_name):
                if cropped:
                    format, imgstr = cropped.split(';base64,')
                    ext = format.split('/')[-1]
                    setattr(product_image, field_name, ContentFile(base64.b64decode(imgstr), name=f'{field_name}_{variant.id}.{ext}'))
                elif raw:
                    setattr(product_image, field_name, raw)

            save_cropped_image(image1_cropped, image1, 'image1')
            save_cropped_image(image2_cropped, image2, 'image2')
            save_cropped_image(image3_cropped, image3, 'image3')

            product_image.save()
            messages.success(request, 'Variants updated successfully!')
            return redirect('books')

    return render(request, 'admin/variant_edit.html', {'variant': variant, 'images': images, 'error': error})




@login_required
def admin_order_details(request, order_id):

    order = get_object_or_404(Order, id=order_id)
    
    logger.info(f"admin side status: {order_id}")

    if request.method == 'POST':
        
        item_id = request.POST.get('item_id')
        new_status = request.POST.get('status')
        
        if not item_id:
            messages.error(request, "Item ID is required.")
            return redirect('admin_order_details', order_id=order.id)
        
        try:
            order_item = OrderItem.objects.get(id=item_id, order=order)
        except OrderItem.DoesNotExist:
            messages.error(request, "Order item not found.")
            return redirect('admin_order_details', order_id=order.id)
        
        logger.info(f"admin side updating item {item_id} status to: {new_status}")
        
        valid_statuses = ['pending', 'shipped', 'delivered', 'cancelled', 'request to cancel','return approved','request to return']
        
        if new_status not in valid_statuses:
            messages.error(request, "Invalid status selected.")
            return redirect('admin_order_details', order_id=order.id)
        
        # Handle status updates per item
        old_status = order_item.status
        refund_message_shown = False
        
        # If item is being delivered, create COD transaction for this specific order item
        if new_status == 'delivered' and order.payment_method == 'cod':
            # Check if COD transaction already exists for this order item
            existing_transaction = Transaction.objects.filter(
                order=order,
                order_item=order_item,
                transaction_type='cod'
            ).first()
            
            if not existing_transaction:
                # Calculate amount for this order item (with proportional coupon discount)
                from decimal import Decimal
                item_gross = Decimal(str(order_item.unit_price)) * order_item.quantity
                item_product_discount = (Decimal(str(order_item.unit_price)) - Decimal(str(order_item.discount_price))) * order_item.quantity
                coupon_discount_per_item = Decimal('0.00')
                if order.coupon_discount and order.coupon_discount > 0:
                    total_items = order.order_items.count()
                    if total_items > 0:
                        coupon_discount_per_item = Decimal(str(order.coupon_discount)) / total_items
                item_net_amount = item_gross - item_product_discount - coupon_discount_per_item
                
                # Create COD transaction for this order item
                Transaction.objects.create(
                    user=order.user,
                    order=order,
                    order_item=order_item,
                    transaction_type='cod',
                    amount=item_net_amount,
                    description=f'COD payment for item {order_item.product_variant.product.book_title} from order {order.order_id}',
                    payment_method='cod',
                    status='completed'
                )
            
            # Mark order as paid if all items are delivered
            all_delivered = order.order_items.exclude(status='delivered').count() == 0
            if all_delivered:
                order.is_paid = True
                order.save()
        
        # Handle cancel/return request approvals with refunds
        if old_status == 'request to return' and new_status == 'return approved':
            # If returning, update stock and refund if paid
            variant = order_item.product_variant
            variant.available_quantity = F('available_quantity') + order_item.quantity
            variant.save(update_fields=['available_quantity'])
            
            if order.is_paid:
                user = order.user
                # Calculate accurate refund amount (gross - product discount - proportional coupon discount)
                from decimal import Decimal
                item_gross = Decimal(str(order_item.unit_price)) * order_item.quantity
                item_product_discount = (Decimal(str(order_item.unit_price)) - Decimal(str(order_item.discount_price))) * order_item.quantity
                coupon_discount_per_item = Decimal('0.00')
                if order.coupon_discount and order.coupon_discount > 0:
                    total_items = order.order_items.count()
                    if total_items > 0:
                        coupon_discount_per_item = Decimal(str(order.coupon_discount)) / total_items
                item_refund_amount = item_gross - item_product_discount - coupon_discount_per_item
                
                user_wallet, created = Wallet.objects.get_or_create(user=user)
                user_wallet.wallet_amount += item_refund_amount
                user_wallet.save()
                try:
                    Transaction.objects.create(
                        user=user,
                        order=order,
                        order_item=order_item,
                        transaction_type='refund',
                        amount=item_refund_amount,
                        description=f'Refund for returned item {order_item.id} ({order_item.product_variant.product.book_title}) from order {order.order_id}',
                        payment_method='wallet',
                        status='completed'
                    )
                except Exception as e:
                    logger.error(f"Failed to log wallet refund: {e}")
                messages.success(request, f"Refunded ₹{item_refund_amount:.2f} to user {user.email}'s wallet for item {order_item.product_variant.product.book_title}.")
                refund_message_shown = True
        
        # If item is being cancelled/returned and order is paid, handle refund (for direct cancellations, not from request)
        elif new_status in ['return approved', 'cancelled'] and order.is_paid and old_status not in ['return approved', 'cancelled', 'request to return', 'request to cancel']:
            user = order.user
            
            # Calculate accurate refund amount (gross - product discount - proportional coupon discount)
            from decimal import Decimal
            item_gross = Decimal(str(order_item.unit_price)) * order_item.quantity
            item_product_discount = (Decimal(str(order_item.unit_price)) - Decimal(str(order_item.discount_price))) * order_item.quantity
            coupon_discount_per_item = Decimal('0.00')
            if order.coupon_discount and order.coupon_discount > 0:
                total_items = order.order_items.count()
                if total_items > 0:
                    coupon_discount_per_item = Decimal(str(order.coupon_discount)) / total_items
            item_refund_amount = item_gross - item_product_discount - coupon_discount_per_item
            
            # Update stock
            variant = order_item.product_variant
            variant.available_quantity = F('available_quantity') + order_item.quantity
            variant.save(update_fields=['available_quantity'])
            
            # Wallet crediting
            user_wallet, created = Wallet.objects.get_or_create(user=user)
            user_wallet.wallet_amount += item_refund_amount
            user_wallet.save()
            
            try:
                Transaction.objects.create(
                    user=user,
                    order=order,
                    order_item=order_item,
                    transaction_type='refund',
                    amount=item_refund_amount,
                    description=f'Refund for order item {order_item.id} ({order_item.product_variant.product.book_title}) from order {order.order_id}',
                    payment_method='wallet',
                    status='completed'
                )
            except Exception as e:
                logger.error(f"Failed to log wallet refund: {e}")
            
            messages.success(request, f"Refunded ₹{item_refund_amount:.2f} to user {user.email}'s wallet for item {order_item.product_variant.product.book_title}.")
            refund_message_shown = True
        
        # Handle cancel request approval
        if old_status == 'request to cancel' and new_status == 'cancelled':
            # Update stock
            variant = order_item.product_variant
            variant.available_quantity = F('available_quantity') + order_item.quantity
            variant.save(update_fields=['available_quantity'])
            
            # If order is paid, refund to wallet
            if order.is_paid and not refund_message_shown:
                user = order.user
                # Calculate accurate refund amount (gross - product discount - proportional coupon discount)
                from decimal import Decimal
                item_gross = Decimal(str(order_item.unit_price)) * order_item.quantity
                item_product_discount = (Decimal(str(order_item.unit_price)) - Decimal(str(order_item.discount_price))) * order_item.quantity
                coupon_discount_per_item = Decimal('0.00')
                if order.coupon_discount and order.coupon_discount > 0:
                    total_items = order.order_items.count()
                    if total_items > 0:
                        coupon_discount_per_item = Decimal(str(order.coupon_discount)) / total_items
                item_refund_amount = item_gross - item_product_discount - coupon_discount_per_item
                
                user_wallet, created = Wallet.objects.get_or_create(user=user)
                user_wallet.wallet_amount += item_refund_amount
                user_wallet.save()
                try:
                    Transaction.objects.create(
                        user=user,
                        order=order,
                        order_item=order_item,
                        transaction_type='refund',
                        amount=item_refund_amount,
                        description=f'Refund for cancelled item {order_item.id} ({order_item.product_variant.product.book_title}) from order {order.order_id}',
                        payment_method='wallet',
                        status='completed'
                    )
                except Exception as e:
                    logger.error(f"Failed to log wallet refund: {e}")
                messages.success(request, f"Refunded ₹{item_refund_amount:.2f} to user {user.email}'s wallet for item {order_item.product_variant.product.book_title}.")
                refund_message_shown = True
        
        # If item is being cancelled directly (not from request) and order is not paid (COD), just update stock
        elif new_status == 'cancelled' and not order.is_paid and old_status != 'cancelled' and old_status != 'request to cancel':
            variant = order_item.product_variant
            variant.available_quantity = F('available_quantity') + order_item.quantity
            variant.save(update_fields=['available_quantity'])
        
        # Update the item status
        order_item.status = new_status
        order_item.save()
        
        # Generate invoice if status is 'shipped' and invoice doesn't exist
        if new_status == 'shipped':
            from users.views import generate_invoice_for_order
            try:
                # Check if Invoice table exists by trying to access it
                from users.models import Invoice
                generate_invoice_for_order(order)
            except Exception as e:
                # Log error but don't fail the status update
                logger.warning(f"Could not generate invoice for order {order.order_id}: {str(e)}. Make sure migrations are run.")
        
        # Only show generic status update message if we haven't shown a refund message
        if not refund_message_shown:
            messages.success(request, f"Order item status updated to {new_status}.")
        
        return redirect('admin_order_details', order_id=order.id)

    # Calculate order summary breakdown
    from decimal import Decimal
    gross_price = Decimal('0.00')
    total_product_discount = Decimal('0.00')
    
    for item in order.order_items.all():
        # Gross price = unit_price * quantity
        item_gross = Decimal(str(item.unit_price)) * item.quantity
        gross_price += item_gross
        
        # Product/Genre discount = (unit_price - discount_price) * quantity
        item_product_discount = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        total_product_discount += item_product_discount
    
    # Coupon discount
    coupon_discount = Decimal('0.00')
    if order.coupon_discount:
        coupon_discount = Decimal(str(order.coupon_discount))
    
    # Total discount = product discount + coupon discount
    total_discount = total_product_discount + coupon_discount
    
    # Grand total = gross price - total discount
    grand_total = gross_price - total_discount
    
    context = {
        'reload': True,
        'order': order,
        'heading': {'name': f'Order Details #{ order.id }'},
        'gross_price': gross_price,
        'total_product_discount': total_product_discount,
        'coupon_discount': coupon_discount,
        'total_discount': total_discount,
        'grand_total': grand_total,
    }

    return render(request, 'admin/admin_order_details.html', context)




def admin_review(request):
    reviews = Review.objects.select_related('user', 'product').order_by('-created_at')
    return render(request, "admin/admin_review.html", {'reviews':reviews})


def toggle_review_status(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    review.is_active = not review.is_active  # Flip True to False or vice versa
    review.save()
    status_text = "Activated" if review.is_active else "Deactivated"
    messages.success(request, f'Review status changed to {status_text}.')
    return redirect('admin_review')


@never_cache
@login_required
def coupon_list(request):
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    discount_type_filter = request.GET.get('discount_type', 'all')
    
    coupons = Coupon.objects.all()
    
    if query:
        coupons = coupons.filter(
            Q(code__icontains=query) |
            Q(description__icontains=query)
        )
    
    if status_filter == 'active':
        coupons = coupons.filter(is_active=True, valid_until__gte=timezone.now())
    elif status_filter == 'inactive':
        coupons = coupons.filter(is_active=False)
    elif status_filter == 'expired':
        coupons = coupons.filter(valid_until__lt=timezone.now())
    
    if discount_type_filter != 'all':
        coupons = coupons.filter(discount_type=discount_type_filter)
    
    coupons = coupons.order_by('-created_at')
    
    # Statistics
    total_coupons = Coupon.objects.count()
    active_coupons = Coupon.objects.filter(is_active=True, valid_until__gte=timezone.now()).count()
    expired_coupons = Coupon.objects.filter(valid_until__lt=timezone.now()).count()
    total_usage = CouponUsage.objects.count()
    
    paginator = Paginator(coupons, 10)
    page = request.GET.get('page')
    
    try:
        coupons = paginator.page(page)
    except PageNotAnInteger:
        coupons = paginator.page(1)
    except EmptyPage:
        coupons = paginator.page(paginator.num_pages)
    
    context = {
        'coupons': coupons,
        'query': query,
        'status_filter': status_filter,
        'discount_type_filter': discount_type_filter,
        'total_coupons': total_coupons,
        'active_coupons': active_coupons,
        'expired_coupons': expired_coupons,
        'total_usage': total_usage,
    }
    
    return render(request, 'coupons/coupon_list.html', context)

@login_required
def add_coupon(request):
    if request.method == 'POST':
        code = request.POST.get('code', '').strip().upper()
        description = request.POST.get('description', '').strip()
        discount_type = request.POST.get('discount_type', 'percentage')
        discount_value = request.POST.get('discount_value', '').strip()
        minimum_amount = request.POST.get('minimum_amount', '0').strip()
        maximum_discount = request.POST.get('maximum_discount', '').strip()
        usage_limit = request.POST.get('usage_limit', '').strip()
        valid_from = request.POST.get('valid_from', '').strip()
        valid_until = request.POST.get('valid_until', '').strip()
        
        errors = []
        
        # Validate coupon code
        if not code:
            errors.append("Coupon code is required")
        elif len(code) < 3:
            errors.append("Coupon code should be at least 3 characters long")
        elif len(code) > 50:
            errors.append("Coupon code should not exceed 50 characters")
        elif not re.match(r'^[A-Z0-9]+$', code):
            errors.append("Coupon code should only contain uppercase letters and numbers")
        elif Coupon.objects.filter(code=code).exists():
            errors.append("A coupon with this code already exists")
        
        # Validate discount value
        if not discount_value:
            errors.append("Discount value is required")
        else:
            try:
                discount_value = float(discount_value)
                if discount_value <= 0:
                    errors.append("Discount value must be greater than 0")
                if discount_type == 'percentage' and discount_value > 100:
                    errors.append("Percentage discount cannot exceed 100%")
            except ValueError:
                errors.append("Please enter a valid discount value")
        
        # Validate minimum amount
        try:
            minimum_amount = float(minimum_amount)
            if minimum_amount < 0:
                errors.append("Minimum amount cannot be negative")
        except ValueError:
            errors.append("Please enter a valid minimum amount")
        
        # Validate maximum discount
        if maximum_discount:
            try:
                maximum_discount = float(maximum_discount)
                if maximum_discount <= 0:
                    errors.append("Maximum discount must be greater than 0")
            except ValueError:
                errors.append("Please enter a valid maximum discount")
        else:
            maximum_discount = None
       
        
        # Validate dates
        if not valid_from:
            errors.append("Valid from date is required")
        if not valid_until:
            errors.append("Valid until date is required")
        
        if valid_from and valid_until:
            try:
                from_date = timezone.make_aware(datetime.strptime(valid_from, '%Y-%m-%dT%H:%M'))
                until_date = timezone.make_aware(datetime.strptime(valid_until, '%Y-%m-%dT%H:%M'))
                
                if from_date >= until_date:
                    errors.append("Valid from date must be before valid until date")
                
                if until_date <= timezone.now():
                    errors.append("Valid until date must be in the future")
                    
            except ValueError:
                errors.append("Please enter valid dates")
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'coupons/add_coupon.html', {
                'code': request.POST.get('code', ''),
                'description': description,
                'discount_type': discount_type,
                'discount_value': request.POST.get('discount_value', ''),
                'minimum_amount': request.POST.get('minimum_amount', ''),
                'maximum_discount': request.POST.get('maximum_discount', ''),
                
                'valid_from': valid_from,
                'valid_until': valid_until,
            })
        
        try:
            Coupon.objects.create(
                code=code,
                description=description,
                discount_type=discount_type,
                discount_value=discount_value,
                minimum_amount=minimum_amount,
                maximum_discount=maximum_discount,
                valid_from=from_date,
                valid_until=until_date,
                is_active=True
            )
            messages.success(request, f"Coupon '{code}' created successfully")
            return redirect('coupon_list')
        except Exception as e:
            messages.error(request, "An error occurred while creating the coupon")
            return render(request, 'coupons/add_coupon.html', {
                'code': request.POST.get('code', ''),
                'description': description,
                'discount_type': discount_type,
                'discount_value': request.POST.get('discount_value', ''),
                'minimum_amount': request.POST.get('minimum_amount', ''),
                'maximum_discount': request.POST.get('maximum_discount', ''),
                'valid_from': valid_from,
                'valid_until': valid_until,
            })
    
    return render(request, 'coupons/add_coupon.html')




@login_required
def edit_coupon(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    
    if request.method == 'POST':
        code = request.POST.get('code', '').strip().upper()
        description = request.POST.get('description', '').strip()
        discount_type = request.POST.get('discount_type', 'percentage')
        discount_value = request.POST.get('discount_value', '').strip()
        minimum_amount = request.POST.get('minimum_amount', '0').strip()
        maximum_discount = request.POST.get('maximum_discount', '').strip()
        usage_limit = request.POST.get('usage_limit', '').strip()
        valid_from = request.POST.get('valid_from', '').strip()
        valid_until = request.POST.get('valid_until', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        errors = []
        
        # Validate coupon code
        if not code:
            errors.append("Coupon code is required")
        elif len(code) < 3:
            errors.append("Coupon code should be at least 3 characters long")
        elif len(code) > 50:
            errors.append("Coupon code should not exceed 50 characters")
        elif not re.match(r'^[A-Z0-9]+$', code):
            errors.append("Coupon code should only contain uppercase letters and numbers")
        elif Coupon.objects.filter(code=code).exclude(id=coupon.id).exists():
            errors.append("A coupon with this code already exists")
        
        # Validate discount value
        if not discount_value:
            errors.append("Discount value is required")
        else:
            try:
                discount_value = float(discount_value)
                if discount_value <= 0:
                    errors.append("Discount value must be greater than 0")
                if discount_type == 'percentage' and discount_value > 100:
                    errors.append("Percentage discount cannot exceed 100%")
            except ValueError:
                errors.append("Please enter a valid discount value")
        
        # Validate minimum amount
        try:
            minimum_amount = float(minimum_amount)
            if minimum_amount < 0:
                errors.append("Minimum amount cannot be negative")
        except ValueError:
            errors.append("Please enter a valid minimum amount")
        
        # Validate maximum discount
        if maximum_discount:
            try:
                maximum_discount = float(maximum_discount)
                if maximum_discount <= 0:
                    errors.append("Maximum discount must be greater than 0")
            except ValueError:
                errors.append("Please enter a valid maximum discount")
        else:
            maximum_discount = None
        
       
        
        # Validate dates
        if not valid_from:
            errors.append("Valid from date is required")
        if not valid_until:
            errors.append("Valid until date is required")
        
        if valid_from and valid_until:
            try:
                from_date = timezone.make_aware(datetime.strptime(valid_from, '%Y-%m-%dT%H:%M'))
                until_date = timezone.make_aware(datetime.strptime(valid_until, '%Y-%m-%dT%H:%M'))
                
                if from_date >= until_date:
                    errors.append("Valid from date must be before valid until date")
                    
            except ValueError:
                errors.append("Please enter valid dates")
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                coupon.code = code
                coupon.description = description
                coupon.discount_type = discount_type
                coupon.discount_value = discount_value
                coupon.minimum_amount = minimum_amount
                coupon.maximum_discount = maximum_discount
                coupon.valid_from = from_date
                coupon.valid_until = until_date
                coupon.is_active = is_active
                coupon.save()
                
                messages.success(request, f"Coupon '{code}' updated successfully")
                return redirect('coupon_list')
            except Exception as e:
                messages.error(request, "An error occurred while updating the coupon")
    
    # Get usage statistics
    usage_count = CouponUsage.objects.filter(coupon=coupon).count()
    
    context = {
        'coupon': coupon,
        'usage_count': usage_count,
    }
    
    return render(request, 'coupons/edit_coupon.html', context)

@login_required
def toggle_coupon_status(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'activate':
            coupon.is_active = True
            coupon.save()
            messages.success(request, f"Coupon '{coupon.code}' has been activated successfully")
        elif action == 'deactivate':
            coupon.is_active = False
            coupon.save()
            messages.success(request, f"Coupon '{coupon.code}' has been deactivated successfully")
        
        return redirect('coupon_list')
    
    # Get usage statistics for confirmation
    usage_count = CouponUsage.objects.filter(coupon=coupon).count()
    
    return render(request, 'coupons/toggle_coupon_status.html', {
        'coupon': coupon,
        'usage_count': usage_count,
    })


# Sales Report Views
def sales_report(request):
    """Generate sales report with filtering options"""
    filter_type = request.GET.get('filter', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    orders = Order.objects.filter(
        status__in=['pending', 'shipped', 'delivered'],
        is_active=True,
        is_paid=True  # Only show paid orders
    ).select_related('user', 'address').prefetch_related('order_items__product_variant__product')
    
    now = timezone.now()
    filter_label = "All Time"
    
    if filter_type == 'today':
        start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
        orders = orders.filter(created_at__gte=start_of_day)
        filter_label = "Today"
    elif filter_type == 'week':
        start_of_week = now - timedelta(days=7)
        orders = orders.filter(created_at__gte=start_of_week)
        filter_label = "Past 7 Days"
    elif filter_type == 'month':
        start_of_month = now - timedelta(days=30)
        orders = orders.filter(created_at__gte=start_of_month)
        filter_label = "Past 30 Days"
    elif filter_type == 'custom' and start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            orders = orders.filter(created_at__range=[start, end])
            filter_label = f"{start_date} to {end_date}"
        except ValueError:
            messages.error(request, "Invalid date format")
    
    # Calculate summary from orders (count unique orders)
    # Get unique order count before aggregating
    unique_orders_count = orders.count()
    
    # Get order items to calculate gross total and product discounts
    from users.models import OrderItem
    order_items_for_summary = OrderItem.objects.filter(order__in=orders)
    
    # Calculate total gross total (sum of unit_price * quantity for all items)
    from django.db.models import F, Sum
    total_gross_total = order_items_for_summary.aggregate(
        total=Sum(F('unit_price') * F('quantity'))
    )['total'] or 0
    
    # Calculate total product discount (sum of (unit_price - discount_price) * quantity)
    total_product_discount = order_items_for_summary.aggregate(
        total=Sum((F('unit_price') - F('discount_price')) * F('quantity'))
    )['total'] or 0
    
    # Calculate total coupon discount from orders
    total_coupon_discount = orders.aggregate(
        total=Sum('coupon_discount')
    )['total'] or 0
    
    # Total deductions = product discount + coupon discount
    total_deductions = total_product_discount + total_coupon_discount
    
    # Calculate total refunds as sum of net amounts of cancelled/return approved order items
    from decimal import Decimal
    # Get cancelled and return approved order items
    refunded_items = OrderItem.objects.filter(
        order__in=orders,
        status__in=['cancelled', 'return approved']
    )
    
    # Calculate total refund by summing net amounts of refunded items
    total_refund = Decimal('0.00')
    for item in refunded_items:
        # Gross Price = unit_price * quantity
        gross = Decimal(str(item.unit_price)) * item.quantity
        
        # Product/Genre Discount = (unit_price - discount_price) * quantity
        product_disc = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        
        # Coupon Discount per item = (order coupon_discount / number of items in order)
        coupon_disc_per_item = Decimal('0.00')
        if item.order.coupon_discount and item.order.coupon_discount > 0:
            total_items = item.order.order_items.count()
            if total_items > 0:
                coupon_disc_per_item = Decimal(str(item.order.coupon_discount)) / total_items
        
        # Total Discount = Product Discount + Coupon Discount per item
        total_disc = product_disc + coupon_disc_per_item
        
        # Net Amount = Gross Price - Discount Price
        net_amt = gross - total_disc
        total_refund += net_amt
    
    # Convert to Decimal for precise calculation
    total_gross_total = Decimal(str(total_gross_total))
    total_deductions = Decimal(str(total_deductions))
    # total_refund is already Decimal from the loop above
    
    # Get order items from filtered orders (OrderItem already imported above)
    order_items = OrderItem.objects.filter(
        order__in=orders
    ).select_related(
        'order', 
        'order__user', 
        'product_variant', 
        'product_variant__product'
    ).prefetch_related('order__order_items').order_by('-order__created_at', '-id')
    
    # Calculate total net amount from all order items (before pagination)
    total_net_amount_from_items = Decimal('0.00')
    for item in order_items:
        # Gross Price = unit_price * quantity
        gross = Decimal(str(item.unit_price)) * item.quantity
        
        # Product/Genre Discount = (unit_price - discount_price) * quantity
        product_disc = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        
        # Coupon Discount per item = (order coupon_discount / number of items in order)
        coupon_disc_per_item = Decimal('0.00')
        if item.order.coupon_discount and item.order.coupon_discount > 0:
            # Use prefetched order_items to avoid extra queries
            total_items = len(item.order.order_items.all())
            if total_items > 0:
                coupon_disc_per_item = Decimal(str(item.order.coupon_discount)) / total_items
        
        # Total Discount = Product Discount + Coupon Discount per item
        total_disc = product_disc + coupon_disc_per_item
        
        # Net Amount = Gross Price - Discount Price
        net_amt = gross - total_disc
        total_net_amount_from_items += net_amt
    
    # Total Net Amount = Sum of all row net amounts - Total Refund
    total_net_amount = total_net_amount_from_items - total_refund
    
    summary = {
        'total_orders': unique_orders_count,
        'total_gross_total': total_gross_total,
        'total_deductions': total_deductions,
        'total_refund': total_refund,
        'total_net_amount': total_net_amount,
    }
    
    # Pagination for order items (after calculating totals)
    paginator = Paginator(order_items, 20)
    page = request.GET.get('page', 1)
    
    try:
        order_items_page = paginator.page(page)
    except PageNotAnInteger:
        order_items_page = paginator.page(1)
    except EmptyPage:
        order_items_page = paginator.page(paginator.num_pages)
    
    # Add calculated fields to each order item in the paginated page for display
    for item in order_items_page:
        # Gross Price = unit_price * quantity
        item.gross_price = Decimal(str(item.unit_price)) * item.quantity
        
        # Product/Genre Discount = (unit_price - discount_price) * quantity
        product_discount = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        
        # Coupon Discount per item = (order coupon_discount / number of items in order)
        coupon_discount_per_item = Decimal('0.00')
        if item.order.coupon_discount and item.order.coupon_discount > 0:
            # Use prefetched order_items to avoid extra queries
            total_items_in_order = len(item.order.order_items.all())
            if total_items_in_order > 0:
                coupon_discount_per_item = Decimal(str(item.order.coupon_discount)) / total_items_in_order
        
        # Total Discount Amount = Product Discount + Coupon Discount per item
        item.discount_amount = product_discount + coupon_discount_per_item
        
        # Net Amount per row = Gross Price - Discount Price
        item.net_amount = item.gross_price - item.discount_amount
    
    context = {
        'order_items': order_items_page,
        'summary': summary,
        'filter_type': filter_type,
        'filter_label': filter_label,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'admin/sales_report.html', context)


def download_sales_report_pdf(request):
    """Download sales report as PDF with order items"""
    # Get same filters as main report
    filter_type = request.GET.get('filter', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Apply same filtering logic as main view
    orders = Order.objects.filter(
        status__in=['pending', 'shipped', 'delivered'],
        is_active=True,
        is_paid=True  # Only show paid orders
    ).select_related('user', 'address').prefetch_related('order_items__product_variant__product')
    
    now = timezone.now()
    filter_label = "All Time"
    
    if filter_type == 'today':
        start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
        orders = orders.filter(created_at__gte=start_of_day)
        filter_label = "Today"
    elif filter_type == 'week':
        start_of_week = now - timedelta(days=7)
        orders = orders.filter(created_at__gte=start_of_week)
        filter_label = "Past 7 Days"
    elif filter_type == 'month':
        start_of_month = now - timedelta(days=30)
        orders = orders.filter(created_at__gte=start_of_month)
        filter_label = "Past 30 Days"
    elif filter_type == 'custom' and start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            orders = orders.filter(created_at__range=[start, end])
            filter_label = f"{start_date} to {end_date}"
        except ValueError:
            pass
    
    # Calculate summary using same logic as main view
    unique_orders_count = orders.count()
    
    from users.models import OrderItem
    from decimal import Decimal
    from django.db.models import F, Sum
    
    order_items_for_summary = OrderItem.objects.filter(order__in=orders)
    
    total_gross_total = order_items_for_summary.aggregate(
        total=Sum(F('unit_price') * F('quantity'))
    )['total'] or 0
    
    total_product_discount = order_items_for_summary.aggregate(
        total=Sum((F('unit_price') - F('discount_price')) * F('quantity'))
    )['total'] or 0
    
    total_coupon_discount = orders.aggregate(
        total=Sum('coupon_discount')
    )['total'] or 0
    
    total_deductions = total_product_discount + total_coupon_discount
    
    # Calculate total refunds from cancelled/return approved items
    refunded_items = OrderItem.objects.filter(
        order__in=orders,
        status__in=['cancelled', 'return approved']
    )
    
    total_refund = Decimal('0.00')
    for item in refunded_items:
        gross = Decimal(str(item.unit_price)) * item.quantity
        product_disc = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        coupon_disc_per_item = Decimal('0.00')
        if item.order.coupon_discount and item.order.coupon_discount > 0:
            total_items = item.order.order_items.count()
            if total_items > 0:
                coupon_disc_per_item = Decimal(str(item.order.coupon_discount)) / total_items
        total_disc = product_disc + coupon_disc_per_item
        net_amt = gross - total_disc
        total_refund += net_amt
    
    # Calculate total net amount
    total_net_amount_from_items = Decimal('0.00')
    order_items_all = OrderItem.objects.filter(order__in=orders).select_related('order', 'product_variant', 'product_variant__product').prefetch_related('order__order_items')
    
    for item in order_items_all:
        gross = Decimal(str(item.unit_price)) * item.quantity
        product_disc = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        coupon_disc_per_item = Decimal('0.00')
        if item.order.coupon_discount and item.order.coupon_discount > 0:
            total_items = len(item.order.order_items.all())
            if total_items > 0:
                coupon_disc_per_item = Decimal(str(item.order.coupon_discount)) / total_items
        total_disc = product_disc + coupon_disc_per_item
        net_amt = gross - total_disc
        total_net_amount_from_items += net_amt
    
    total_net_amount = total_net_amount_from_items - total_refund
    
    # Convert to Decimal
    total_gross_total = Decimal(str(total_gross_total))
    total_deductions = Decimal(str(total_deductions))
    
    # Create PDF with wider margins to accommodate table
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph("Book Hive - Sales Report", title_style))
    elements.append(Paragraph(f"Period: {filter_label}", styles['Normal']))
    elements.append(Paragraph(f"Generated: {now.strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary table
    summary_data = [
        ['Metric', 'Value'],
        ['Total Orders', str(unique_orders_count)],
        ['Total Gross Total', f"Rs.{float(total_gross_total):.2f}"],
        ['Total Deductions', f"Rs.{float(total_deductions):.2f}"],
        ['Total Refund', f"Rs.{float(total_refund):.2f}"],
        ['Total Net Amount', f"Rs.{float(total_net_amount):.2f}"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Order Items table
    elements.append(Paragraph("Order Items Details", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    # Get order items with calculated fields
    order_items = OrderItem.objects.filter(
        order__in=orders
    ).select_related(
        'order', 'order__user', 'product_variant', 'product_variant__product'
    ).prefetch_related('order__order_items').order_by('-order__created_at', '-id')[:200]  # Limit to 200 items for PDF
    
    # Calculate fields for each item
    for item in order_items:
        item.gross_price = Decimal(str(item.unit_price)) * item.quantity
        product_discount = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        coupon_discount_per_item = Decimal('0.00')
        if item.order.coupon_discount and item.order.coupon_discount > 0:
            total_items = len(item.order.order_items.all())
            if total_items > 0:
                coupon_discount_per_item = Decimal(str(item.order.coupon_discount)) / total_items
        item.discount_amount = product_discount + coupon_discount_per_item
        item.net_amount = item.gross_price - item.discount_amount
    
    # Create styles for table cells with text wrapping
    cell_style_normal = ParagraphStyle(
        'CellNormal',
        parent=styles['Normal'],
        fontSize=7,
        leading=8.5,
        wordWrap='CJK',  # Enable word wrapping for all languages
        spaceAfter=1,
        spaceBefore=1,
        leftIndent=2,
        rightIndent=2,
    )
    
    cell_style_header = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontSize=7.5,
        leading=9,
        fontName='Helvetica-Bold',
        textColor=colors.whitesmoke,
        wordWrap='CJK',
        spaceAfter=1,
        spaceBefore=1,
        leftIndent=2,
        rightIndent=2,
    )
    
    # Special style for narrow columns (Qty, Status, Coupon) - smaller font
    cell_style_narrow = ParagraphStyle(
        'CellNarrow',
        parent=styles['Normal'],
        fontSize=6.5,
        leading=8,
        wordWrap='CJK',
        spaceAfter=1,
        spaceBefore=1,
        leftIndent=2,
        rightIndent=2,
        alignment=1,  # Center alignment
    )
    
    # Build table data with Paragraph objects for text wrapping
    order_items_data = []
    
    # Header row with Paragraph objects
    header_row = [
        Paragraph('Order ID', cell_style_header),
        Paragraph('Date', cell_style_header),
        Paragraph('Customer', cell_style_header),
        Paragraph('Product', cell_style_header),
        Paragraph('Qty', cell_style_narrow),
        Paragraph('Gross Price', cell_style_header),
        Paragraph('Discount', cell_style_header),
        Paragraph('Net Amount', cell_style_header),
        Paragraph('Coupon', cell_style_narrow),
        Paragraph('Status', cell_style_narrow),
    ]
    order_items_data.append(header_row)
    
    # Data rows with Paragraph objects
    for item in order_items:
        customer_name = f"{item.order.user.first_name} {item.order.user.last_name}"
        product_name = item.product_variant.product.book_title  # Full name, will wrap
        coupon_code = item.order.coupon_code or 'N/A'
        status_text = item.status.title()
        
        order_items_data.append([
            Paragraph(str(item.order.order_id), cell_style_normal),
            Paragraph(item.order.created_at.strftime('%d-%m-%Y'), cell_style_normal),
            Paragraph(customer_name, cell_style_normal),
            Paragraph(product_name, cell_style_normal),
            Paragraph(str(item.quantity), cell_style_narrow),
            Paragraph(f"Rs.{float(item.gross_price):.2f}", cell_style_normal),
            Paragraph(f"Rs.{float(item.discount_amount):.2f}", cell_style_normal),
            Paragraph(f"Rs.{float(item.net_amount):.2f}", cell_style_normal),
            Paragraph(coupon_code, cell_style_narrow),
            Paragraph(status_text, cell_style_narrow),
        ])
    
    # Adjust column widths for order items table - ensure all content fits on A4 page
    # A4 width = 8.27 inches, minus margins (0.2 + 0.2 = 0.4) = 7.87 inches available
    # Distribute widths to ensure all columns are visible and text wraps properly
    col_widths = [
        0.95*inch,  # Order ID - enough for longer IDs
        0.7*inch,   # Date
        0.95*inch,  # Customer - enough for full names
        1.5*inch,   # Product - increased for long book titles (will wrap)
        0.4*inch,   # Qty
        0.8*inch,    # Gross Price
        0.8*inch,    # Discount
        0.8*inch,    # Net Amount
        0.65*inch,  # Coupon
        0.65*inch,   # Status
    ]
    # Total: 7.8 inches (fits within 7.87 inches available with small buffer)
    order_items_table = Table(order_items_data, colWidths=col_widths, repeatRows=1)  # repeatRows=1 to repeat header on each page
    order_items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (3, -1), 'LEFT'),
        ('ALIGN', (4, 0), (4, -1), 'CENTER'),  # Qty
        ('ALIGN', (5, 0), (7, -1), 'RIGHT'),  # Price columns
        ('ALIGN', (8, 0), (9, -1), 'CENTER'),  # Coupon, Status
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7.5),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # TOP alignment for better wrapping
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    elements.append(order_items_table)
    
    # Build PDF
    doc.build(elements)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{filter_label.replace(" ", "_")}.pdf"'
    
    return response


def download_sales_report_excel(request):
    """Download sales report as Excel with order items"""
    # Get same filters as main report
    filter_type = request.GET.get('filter', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Apply same filtering logic as main view
    orders = Order.objects.filter(
        status__in=['pending', 'shipped', 'delivered'],
        is_active=True,
        is_paid=True  # Only show paid orders
    ).select_related('user', 'address').prefetch_related('order_items__product_variant__product')
    
    now = timezone.now()
    filter_label = "All Time"
    
    if filter_type == 'today':
        start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
        orders = orders.filter(created_at__gte=start_of_day)
        filter_label = "Today"
    elif filter_type == 'week':
        start_of_week = now - timedelta(days=7)
        orders = orders.filter(created_at__gte=start_of_week)
        filter_label = "Past 7 Days"
    elif filter_type == 'month':
        start_of_month = now - timedelta(days=30)
        orders = orders.filter(created_at__gte=start_of_month)
        filter_label = "Past 30 Days"
    elif filter_type == 'custom' and start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = end.replace(hour=23, minute=59, second=59)
            orders = orders.filter(created_at__range=[start, end])
            filter_label = f"{start_date} to {end_date}"
        except ValueError:
            pass
    
    # Calculate summary using same logic as main view
    unique_orders_count = orders.count()
    
    from users.models import OrderItem
    from decimal import Decimal
    from django.db.models import F, Sum
    
    order_items_for_summary = OrderItem.objects.filter(order__in=orders)
    
    total_gross_total = order_items_for_summary.aggregate(
        total=Sum(F('unit_price') * F('quantity'))
    )['total'] or 0
    
    total_product_discount = order_items_for_summary.aggregate(
        total=Sum((F('unit_price') - F('discount_price')) * F('quantity'))
    )['total'] or 0
    
    total_coupon_discount = orders.aggregate(
        total=Sum('coupon_discount')
    )['total'] or 0
    
    total_deductions = total_product_discount + total_coupon_discount
    
    # Calculate total refunds from cancelled/return approved items
    refunded_items = OrderItem.objects.filter(
        order__in=orders,
        status__in=['cancelled', 'return approved']
    )
    
    total_refund = Decimal('0.00')
    for item in refunded_items:
        gross = Decimal(str(item.unit_price)) * item.quantity
        product_disc = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        coupon_disc_per_item = Decimal('0.00')
        if item.order.coupon_discount and item.order.coupon_discount > 0:
            total_items = item.order.order_items.count()
            if total_items > 0:
                coupon_disc_per_item = Decimal(str(item.order.coupon_discount)) / total_items
        total_disc = product_disc + coupon_disc_per_item
        net_amt = gross - total_disc
        total_refund += net_amt
    
    # Calculate total net amount
    total_net_amount_from_items = Decimal('0.00')
    order_items_all = OrderItem.objects.filter(order__in=orders).select_related('order', 'product_variant', 'product_variant__product').prefetch_related('order__order_items')
    
    for item in order_items_all:
        gross = Decimal(str(item.unit_price)) * item.quantity
        product_disc = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        coupon_disc_per_item = Decimal('0.00')
        if item.order.coupon_discount and item.order.coupon_discount > 0:
            total_items = len(item.order.order_items.all())
            if total_items > 0:
                coupon_disc_per_item = Decimal(str(item.order.coupon_discount)) / total_items
        total_disc = product_disc + coupon_disc_per_item
        net_amt = gross - total_disc
        total_net_amount_from_items += net_amt
    
    total_net_amount = total_net_amount_from_items - total_refund
    
    # Convert to Decimal
    total_gross_total = Decimal(str(total_gross_total))
    total_deductions = Decimal(str(total_deductions))
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Styling
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Book Hive - Sales Report"
    ws['A1'].font = Font(bold=True, size=16, color="2563eb")
    ws.merge_cells('A1:J1')
    
    ws['A2'] = f"Period: {filter_label}"
    ws['A3'] = f"Generated: {now.strftime('%Y-%m-%d %H:%M')}"
    
    # Summary section
    ws['A5'] = "Summary"
    ws['A5'].font = Font(bold=True, size=14)
    
    summary_headers = ['Metric', 'Value']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    summary_data = [
        ['Total Orders', unique_orders_count],
        ['Total Gross Total', f"Rs.{float(total_gross_total):.2f}"],
        ['Total Deductions', f"Rs.{float(total_deductions):.2f}"],
        ['Total Refund', f"Rs.{float(total_refund):.2f}"],
        ['Total Net Amount', f"Rs.{float(total_net_amount):.2f}"],
    ]
    
    for row_idx, (metric, value) in enumerate(summary_data, 7):
        ws.cell(row=row_idx, column=1, value=metric).border = border
        ws.cell(row=row_idx, column=2, value=value).border = border
    
    # Order Items section
    ws['A13'] = "Order Items Details"
    ws['A13'].font = Font(bold=True, size=14)
    
    # Headers for order items
    headers = ['Order ID', 'Date', 'Customer', 'Product', 'Quantity', 'Gross Price', 'Discount Price', 'Net Amount', 'Coupon', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=14, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get order items with calculated fields
    order_items = OrderItem.objects.filter(
        order__in=orders
    ).select_related(
        'order', 'order__user', 'product_variant', 'product_variant__product'
    ).prefetch_related('order__order_items').order_by('-order__created_at', '-id')
    
    # Calculate fields for each item
    for item in order_items:
        item.gross_price = Decimal(str(item.unit_price)) * item.quantity
        product_discount = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        coupon_discount_per_item = Decimal('0.00')
        if item.order.coupon_discount and item.order.coupon_discount > 0:
            total_items = len(item.order.order_items.all())
            if total_items > 0:
                coupon_discount_per_item = Decimal(str(item.order.coupon_discount)) / total_items
        item.discount_amount = product_discount + coupon_discount_per_item
        item.net_amount = item.gross_price - item.discount_amount
    
    # Order items data
    row_num = 15
    for item in order_items:
        ws.cell(row=row_num, column=1, value=item.order.order_id).border = border
        ws.cell(row=row_num, column=2, value=item.order.created_at.strftime('%Y-%m-%d %H:%M')).border = border
        ws.cell(row=row_num, column=3, value=f"{item.order.user.first_name} {item.order.user.last_name}").border = border
        ws.cell(row=row_num, column=4, value=item.product_variant.product.book_title).border = border
        ws.cell(row=row_num, column=5, value=item.quantity).border = border
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=6, value=float(item.gross_price)).border = border
        ws.cell(row=row_num, column=6).number_format = 'Rs.#,##0.00'
        ws.cell(row=row_num, column=7, value=float(item.discount_amount)).border = border
        ws.cell(row=row_num, column=7).number_format = 'Rs.#,##0.00'
        ws.cell(row=row_num, column=8, value=float(item.net_amount)).border = border
        ws.cell(row=row_num, column=8).number_format = 'Rs.#,##0.00'
        ws.cell(row=row_num, column=9, value=item.order.coupon_code or 'N/A').border = border
        ws.cell(row=row_num, column=9).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=10, value=item.status.title()).border = border
        ws.cell(row=row_num, column=10).alignment = Alignment(horizontal='center')
        row_num += 1
    
    # Adjust column widths
    column_widths = {
        'A': 15,  # Order ID
        'B': 18,  # Date
        'C': 20,  # Customer
        'D': 35,  # Product
        'E': 10,  # Quantity
        'F': 12,  # Gross Price
        'G': 12,  # Discount Price
        'H': 12,  # Net Amount
        'I': 12,  # Coupon
        'J': 12,  # Status
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_report_{filter_label.replace(" ", "_")}.xlsx"'
    
    wb.save(response)
    return response
@login_required
def admin_customer_queries(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    queries = CustomerSupport.objects.filter(user=user).order_by('-created_at')
    
    context = {
        'user': user,
        'queries': queries,
    }
    return render(request, 'admin/customer_queries.html', context)

@login_required
def admin_query_reply(request, query_id):
    query = get_object_or_404(CustomerSupport, id=query_id)
    
    if request.method == 'POST':
        reply_message = request.POST.get('reply_message')
        mark_resolved = request.POST.get('mark_resolved') == 'on'
        
        if not reply_message:
            messages.error(request, 'Please enter a reply message.')
            return redirect('admin_query_details', query_id=query_id)
        
        try:
            # Send email to user
            subject = f'Re: {query.subject}'
            email_message = f"""
Dear {query.user.first_name} {query.user.last_name},

Thank you for contacting us. Here is our response to your query:

Your Query:
{query.message}

Our Reply:
{reply_message}

If you have any further questions, please don't hesitate to contact us.

Best regards,
BookHive Support Team
"""
            from_email = settings.DEFAULT_FROM_EMAIL
            send_mail(subject, email_message, from_email, [query.user.email])
            
            # Update query
            query.admin_reply = reply_message
            query.replied_at = timezone.now()
            if mark_resolved:
                query.status = 'resolved'
            query.save()
            
            messages.success(request, f'Reply sent successfully to {query.user.email}')
            return redirect('admin_query_details', query_id=query_id)
            
        except Exception as e:
            messages.error(request, f'Error sending reply: {str(e)}')
            return redirect('admin_query_details', query_id=query_id)
    
    return redirect('admin_query_details', query_id=query_id)

@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
def admin_transactions(request):
    """Display list of all transactions including order item level transactions"""
    # Get all transactions with order_item information
    transactions = Transaction.objects.select_related(
        'user', 'order', 'order_item', 'order_item__product_variant', 'order_item__product_variant__product'
    ).all().order_by('-transaction_date')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        transactions = transactions.filter(
            Q(transaction_id__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(order__order_id__icontains=search_query) |
            Q(order_item__product_variant__product__book_title__icontains=search_query)
        )
    
    # Filter by transaction type
    filter_method = request.GET.get('filter_method', '')
    if filter_method:
        # Map display name to transaction type value
        type_mapping = {
            'razorpay': 'razorpay',
            'cod': 'cod',
            'refund': 'refund',
            'wallet addition': 'wallet_addition',
            'wallet debit': 'wallet_debit',
            'wallet payment': 'wallet_debit',
        }
        trans_type = type_mapping.get(filter_method.lower(), '')
        if trans_type:
            transactions = transactions.filter(transaction_type=trans_type)
        else:
            # Also check payment method for backward compatibility
            transactions = transactions.filter(payment_method__icontains=filter_method)
    
    # Get unique transaction types for filter dropdown
    payment_methods = []
    for choice_value, choice_display in Transaction.TRANSACTION_TYPE_CHOICES:
        payment_methods.append(choice_display)
    
    # Pagination
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(transactions, 20)
    page = request.GET.get('page', 1)
    try:
        transactions_page = paginator.page(page)
    except PageNotAnInteger:
        transactions_page = paginator.page(1)
    except EmptyPage:
        transactions_page = paginator.page(paginator.num_pages)
    
    context = {
        'transactions': transactions_page,
        'payment_methods': sorted(set(payment_methods)),
        'search_query': search_query,
        'filter_method': filter_method,
    }
    
    return render(request, 'admin/admin_transactions.html', context)


@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
def transaction_detail(request, order_id):
    """Display detailed view of a specific transaction"""
    # Find transaction by transaction_id
    transaction = get_object_or_404(
        Transaction.objects.select_related('user', 'order', 'order__address', 'order_item', 'order_item__product_variant', 'order_item__product_variant__product'),
        transaction_id=order_id
    )
    
    context = {
        'transaction': transaction,
    }
    
    # If transaction has an associated order item, calculate pricing breakdown
    if transaction.order_item:
        from decimal import Decimal
        item = transaction.order_item
        
        # Gross price (unit price * quantity)
        gross_price = Decimal(str(item.unit_price)) * item.quantity
        
        # Product/Genre discount (unit_price - discount_price) * quantity
        product_discount = (Decimal(str(item.unit_price)) - Decimal(str(item.discount_price))) * item.quantity
        
        # Coupon discount per item
        coupon_discount_per_item = Decimal('0.00')
        if transaction.order and transaction.order.coupon_discount and transaction.order.coupon_discount > 0:
            total_items = transaction.order.order_items.count()
            if total_items > 0:
                coupon_discount_per_item = Decimal(str(transaction.order.coupon_discount)) / total_items
        
        # Total discount = product discount + coupon discount
        total_discount = product_discount + coupon_discount_per_item
        
        # Net amount (should match transaction.amount)
        net_amount = gross_price - total_discount
        
        context['item_pricing'] = {
            'gross_price': gross_price,
            'product_discount': product_discount,
            'coupon_discount': coupon_discount_per_item,
            'total_discount': total_discount,
            'net_amount': net_amount,
        }
    
    return render(request, 'admin/admin_transaction_detail.html', context)


@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
@login_required(login_url='admin_signin')
def wallet_management(request):
    """Display all wallet transactions with user filtering"""
    
    # Get all wallet transactions (transactions with wallet payment method)
    wallet_transactions = Transaction.objects.filter(
        payment_method='wallet'
    ).select_related('user', 'order').order_by('-transaction_date')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        wallet_transactions = wallet_transactions.filter(
            Q(transaction_id__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__phone_no__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(wallet_transactions, 20)
    page = request.GET.get('page', 1)
    try:
        transactions_page = paginator.page(page)
    except PageNotAnInteger:
        transactions_page = paginator.page(1)
    except EmptyPage:
        transactions_page = paginator.page(paginator.num_pages)
    
    context = {
        'transactions': transactions_page,
        'search_query': search_query,
    }
    
    return render(request, 'admin/wallet_management.html', context)


@never_cache
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
@login_required(login_url='admin_signin')
def wallet_user_details(request, user_id):
    """Display detailed wallet information for a specific user"""
    
    user = get_object_or_404(CustomUser, id=user_id)
    
    # Get user wallet
    try:
        wallet = Wallet.objects.get(user=user)
    except Wallet.DoesNotExist:
        wallet = Wallet.objects.create(user=user, wallet_amount=0)
    
    # Get all wallet-related transactions for this user (wallet payments, refunds, wallet debits)
    wallet_transactions = Transaction.objects.filter(
        user=user
    ).filter(
        Q(transaction_type='wallet_debit') | 
        Q(transaction_type='refund') |
        Q(payment_method='wallet')
    ).select_related('order').order_by('-transaction_date')
    
    # Calculate statistics
    total_transactions = wallet_transactions.count()
    total_credits = wallet_transactions.filter(transaction_type='refund').aggregate(
        total=Sum('amount'))['total'] or 0
    total_debits = wallet_transactions.exclude(transaction_type='refund').aggregate(
        total=Sum('amount'))['total'] or 0
    
    context = {
        'user': user,
        'wallet': wallet,
        'wallet_transactions': wallet_transactions,
        'total_transactions': total_transactions,
        'total_credits': total_credits,
        'total_debits': total_debits,
    }
    
    return render(request, 'admin/wallet_user_details.html', context)


@login_required
def admin_query_details(request, query_id):
    query = get_object_or_404(CustomerSupport, id=query_id)
    
    context = {
        'query': query,
    }
    return render(request, 'admin/query_details.html', context)